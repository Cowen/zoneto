"""Reference dataset downloads and management."""

from __future__ import annotations

import csv
import json
import logging
import shutil
import tempfile
import zipfile
from pathlib import Path

import httpx
import openpyxl

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Reference dataset URLs
# ---------------------------------------------------------------------------
_ZONING_URL = (
    "https://ckan0.cf.opendata.inter.prod-toronto.ca"
    "/dataset/34927e44-fc11-4336-a8aa-a0dfb27658b7"
    "/resource/d75fa1ed-cd04-4a0b-bb6d-2b928ffffa6e"
    "/download/zoning-area-4326.geojson"
)
_HERITAGE_REGISTER_URL = (
    "https://ckan0.cf.opendata.inter.prod-toronto.ca"
    "/dataset/e41da515-5ad1-4bc3-85ea-18ec9e55cd33"
    "/resource/108b1080-d048-439f-a9e8-e8d6cd81bddb"
    "/download/heritage_register_address_points_wgs84.zip"
)
_HERITAGE_DISTRICTS_URL = (
    "https://ckan0.cf.opendata.inter.prod-toronto.ca"
    "/dataset/37a3c911-0813-4e87-90ed-3b9fa6156a63"
    "/resource/8e6b9347-63a8-4dac-91fb-a6491a8c1e5a"
    "/download/heritageconservationdistrict.zip"
)
_SECONDARY_PLANS_URL = (
    "https://ckan0.cf.opendata.inter.prod-toronto.ca"
    "/dataset/70a544e9-ee83-43a4-b0be-0dc973627ad7"
    "/resource/08099a8c-a598-4ca3-8395-e4159cc1ec1a"
    "/download/secondary-plans-data-2017-4326.geojson"
)
_ZONING_HEIGHT_URL = (
    "https://ckan0.cf.opendata.inter.prod-toronto.ca"
    "/dataset/34927e44-fc11-4336-a8aa-a0dfb27658b7"
    "/resource/eec27e60-7c2d-4c46-8fa1-b64f441bcc39"
    "/download/zoning-height-overlay-4326.geojson"
)
_MTSA_URL = (
    "https://ckan0.cf.opendata.inter.prod-toronto.ca"
    "/dataset/f7128f82-810d-4677-8166-8892f51969d3"
    "/resource/2b66ca26-b345-4e62-8568-dcd2cd6c3f91"
    "/download/majortransitstationareadelinations_jan2026.shp.zip"
)
_WARD_CENSUS_URL = (
    "https://ckan0.cf.opendata.inter.prod-toronto.ca"
    "/dataset/6678e1a6-d25f-4dff-b2b7-aa8f042bc2eb"
    "/resource/16a31e1d-b4d9-4cf0-b5b3-2e3937cb4121"
    "/download/2023-wardprofiles-2011-2021-censusdata_rev0719.xlsx"
)
_WARD_GEO_URL = (
    "https://ckan0.cf.opendata.inter.prod-toronto.ca"
    "/dataset/6678e1a6-d25f-4dff-b2b7-aa8f042bc2eb"
    "/resource/9398da69-0622-4eb1-a125-4f8c7f1016a4"
    "/download/2023-wardprofiles-geographicareas.xlsx"
)
# Section 37 Community Benefits — Toronto Open Data CKAN
_SECTION37_CSV_URL = (
    "https://ckan0.cf.opendata.inter.prod-toronto.ca"
    "/dataset/8c0b0109-73fc-4eb5-b849-337db81aadc1"
    "/resource/d52af681-04b2-4b26-9c51-6682e59499b7"
    "/download/secured-by-section-37.csv"
)
# TRCA Regulated Area — ArcGIS FeatureServer (6000+ polygon records, paginated)
# Service confirmed via TRCA Open Data Portal (trca-camaps.opendata.arcgis.com)
_TRCA_FEATURESERVER_URL = (
    "https://services1.arcgis.com/d0ZCwU7eGKVeNiEE/arcgis/rest/services"
    "/TRCA_Regulation_Limit2/FeatureServer/0/query"
)
# Ontario Greenbelt Outer Boundary — Ministry of Natural Resources Shapefile ZIP
# CRS: EPSG:4269 (NAD83), functionally identical to WGS84 for spatial joins
_GREENBELT_URL = "https://ws.gisetl.lrc.gov.on.ca/fmedatadownload/Packages/GBOUTBND.zip"


def _download(url: str, dest: Path) -> None:
    """Download *url* to *dest* (binary)."""
    with httpx.Client(follow_redirects=True, timeout=120) as client:
        r = client.get(url)
        r.raise_for_status()
        dest.write_bytes(r.content)


def _fetch_ward_profiles_csv(ref: Path) -> None:
    """Download ward profile XLSXs, compute metrics, write ward_profiles.csv.

    Output: ward_number,ward_pct_renters,ward_median_income,
    ward_pop_density,ward_pct_detached (one row per ward, wards 1–25).
    """
    census_xlsx = Path(tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name)
    geo_xlsx = Path(tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name)
    try:
        with httpx.Client(follow_redirects=True, timeout=120) as client:
            census_xlsx.write_bytes(
                client.get(_WARD_CENSUS_URL).raise_for_status().content
            )
            geo_xlsx.write_bytes(client.get(_WARD_GEO_URL).raise_for_status().content)

        # --- parse census XLSX ---
        wb = openpyxl.load_workbook(census_xlsx, read_only=True, data_only=True)
        ws = wb["2021 One Variable"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        # Row index 17: header (None, 'Toronto', 'Ward 1', ..., 'Ward 25')
        header = rows[17]
        # Build index: ward_number (int) → column index in rows
        ward_col_idx: dict[int, int] = {}
        for col_i, val in enumerate(header):
            if val and str(val).startswith("Ward "):
                try:
                    ward_col_idx[int(str(val).replace("Ward ", "").strip())] = col_i
                except ValueError:
                    pass

        def _cell(row_idx: int, ward_num: int) -> float | None:
            col_i = ward_col_idx.get(ward_num)
            if col_i is None:
                return None
            v = rows[row_idx][col_i]
            return float(v) if v is not None else None

        # Row 18: total population ('Total - Age', toronto_total, ward1, ...)
        # Row 43: total occupied dwellings
        # Row 44: single-detached dwellings
        # Row 1384: median household income
        # Row 1390: tenant households
        # Row 1394: owner households

        # --- parse geographic areas XLSX for ward area (sq km) ---
        wb2 = openpyxl.load_workbook(geo_xlsx, read_only=True, data_only=True)
        ws2 = wb2[wb2.sheetnames[0]]
        geo_rows = list(ws2.iter_rows(values_only=True))
        wb2.close()

        # Row 11: header; rows 12+ are (ward_num, area_sq_km, ...)
        ward_area: dict[int, float] = {}
        for geo_row in geo_rows[12:]:
            if geo_row[0] is None:
                break
            try:
                ward_area[int(geo_row[0])] = float(geo_row[1])
            except (TypeError, ValueError):
                pass

        # Write output CSV
        out_path = ref / "ward_profiles.csv"
        with out_path.open("w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "ward_number",
                    "ward_pct_renters",
                    "ward_median_income",
                    "ward_pop_density",
                    "ward_pct_detached",
                ]
            )
            for ward_num in sorted(ward_col_idx.keys()):
                population = _cell(18, ward_num)
                total_dwellings = _cell(43, ward_num)
                single_detached = _cell(44, ward_num)
                median_income = _cell(1384, ward_num)
                tenant = _cell(1390, ward_num)
                owner = _cell(1394, ward_num)

                pct_renters: float | None = None
                if tenant is not None and owner is not None and (tenant + owner) > 0:
                    pct_renters = tenant / (tenant + owner) * 100.0

                pop_density: float | None = None
                area = ward_area.get(ward_num)
                if population is not None and area is not None and area > 0:
                    pop_density = population / area

                pct_detached: float | None = None
                if (
                    single_detached is not None
                    and total_dwellings is not None
                    and total_dwellings > 0
                ):
                    pct_detached = single_detached / total_dwellings * 100.0

                writer.writerow(
                    [ward_num, pct_renters, median_income, pop_density, pct_detached]
                )
    finally:
        census_xlsx.unlink(missing_ok=True)
        geo_xlsx.unlink(missing_ok=True)


def _fetch_section37(ref: Path) -> None:
    """Download Toronto Section 37 Community Benefits CSV and write as parquet.

    Source schema: _id, Ward, Council/OMB Date, Monetary Value, ByLaw,
    Location, Community Benefits. Normalizes column names to snake_case.
    Writes to ref/section37.parquet.
    """
    import io  # noqa: PLC0415

    import polars as pl  # noqa: PLC0415

    with httpx.Client(follow_redirects=True, timeout=120) as client:
        r = client.get(_SECTION37_CSV_URL)
        r.raise_for_status()
        content = r.content

    df = pl.read_csv(
        io.BytesIO(content),
        infer_schema_length=10000,
        null_values=["", "N/A", "None"],
        schema_overrides={"Ward": pl.String},
    )

    # Normalize column names: lowercase, spaces → underscore, strip special chars
    rename: dict[str, str] = {}
    for col in df.columns:
        normalized = col.lower().replace("/", "_").replace(" ", "_").strip("_")
        if normalized != col:
            rename[col] = normalized
    if rename:
        df = df.rename(rename)

    # Ensure monetary_value is Float64
    if "monetary_value" in df.columns:
        df = df.with_columns(pl.col("monetary_value").cast(pl.Float64, strict=False))

    # Normalize council_omb_date → council_date (simpler key for joins)
    if "council_omb_date" in df.columns and "council_date" not in df.columns:
        df = df.rename({"council_omb_date": "council_date"})

    # Parse council_date to Date — source strings may include a time component
    if "council_date" in df.columns:
        df = df.with_columns(
            pl.col("council_date")
            .str.slice(0, 10)
            .str.to_date(format="%Y-%m-%d", strict=False)
        )

    dest_path = ref / "section37.parquet"
    df.write_parquet(dest_path)
    logger.info("_fetch_section37: wrote %d rows to %s", len(df), dest_path)


def _fetch_trca_regulated_areas(dest: Path) -> None:
    """Download TRCA regulated areas from ArcGIS FeatureServer into a GeoJSON file.

    Paginates through all records (max 2000 per page) and writes a single
    merged FeatureCollection to *dest*.
    """
    features: list[dict] = []
    page_size = 2000
    offset = 0

    with httpx.Client(follow_redirects=True, timeout=120) as client:
        while True:
            params = {
                "where": "OBJECTID > 0",
                "outFields": "OBJECTID,criteria_layers_contribution",
                "f": "geojson",
                "resultRecordCount": str(page_size),
                "resultOffset": str(offset),
            }
            r = client.get(_TRCA_FEATURESERVER_URL, params=params)
            r.raise_for_status()
            page = r.json()
            page_features = page.get("features", [])
            features.extend(page_features)
            logger.info(
                "_fetch_trca_regulated_areas: fetched %d features (offset %d)",
                len(features),
                offset,
            )
            exceeded = page.get("properties", {}).get("exceededTransferLimit", False)
            if not exceeded or len(page_features) < page_size:
                break
            offset += page_size

    geojson = {"type": "FeatureCollection", "features": features}
    dest.write_text(json.dumps(geojson))
    logger.info(
        "_fetch_trca_regulated_areas: wrote %d features to %s", len(features), dest
    )


def _fetch_greenbelt(dest: Path) -> None:
    """Download Ontario Greenbelt boundary Shapefile ZIP and convert to GeoJSON.

    The ZIP contains a Shapefile in EPSG:4269 (NAD83), which is functionally
    equivalent to WGS84 for spatial joins within Canada. DuckDB spatial reads
    the SHP directly and exports WGS84 GeoJSON — no fiona dependency needed.
    """
    import duckdb  # noqa: PLC0415

    tmp_zip = dest.parent / "_greenbelt_tmp.zip"
    tmp_dir = dest.parent / "_greenbelt_tmp"
    try:
        with httpx.Client(follow_redirects=True, timeout=300) as client:
            r = client.get(_GREENBELT_URL)
            r.raise_for_status()
            tmp_zip.write_bytes(r.content)

        tmp_dir.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(tmp_zip) as zf:
            zf.extractall(tmp_dir)

        shp_files = list(tmp_dir.rglob("*.shp"))
        if not shp_files:
            logger.warning("_fetch_greenbelt: no .shp found in downloaded ZIP")
            return

        shp_path = shp_files[0]
        escaped = str(shp_path).replace("'", "''")

        con = duckdb.connect()
        con.execute("INSTALL spatial; LOAD spatial;")
        # NAD83 (EPSG:4269) ≈ WGS84 at sub-metre precision; treat as 4326 for joins
        rows = con.execute(f"""
            SELECT ST_AsGeoJSON(geom) AS geom_json
            FROM ST_Read('{escaped}')
        """).fetchall()
        con.close()

        features = [
            {"type": "Feature", "geometry": json.loads(row[0]), "properties": {}}
            for row in rows
            if row[0] is not None
        ]
        geojson = {"type": "FeatureCollection", "features": features}
        dest.write_text(json.dumps(geojson))
        logger.info("_fetch_greenbelt: wrote %d features to %s", len(features), dest)
    finally:
        tmp_zip.unlink(missing_ok=True)
        if tmp_dir.exists():
            shutil.rmtree(tmp_dir)


def fetch_reference(data_dir: Path = Path("data")) -> None:
    """Download all reference datasets to *data_dir*/reference/.

    Idempotent: skips files that already exist.
    """
    ref = data_dir / "reference"
    ref.mkdir(parents=True, exist_ok=True)

    # Zoning GeoJSON (full-city, WGS84)
    zoning_geojson = ref / "zoning.geojson"
    if not zoning_geojson.exists():
        _download(_ZONING_URL, zoning_geojson)

    # Zoning height overlay GeoJSON (max storeys/height per area, WGS84)
    zoning_height = ref / "zoning_height.geojson"
    if not zoning_height.exists():
        _download(_ZONING_HEIGHT_URL, zoning_height)

    # Heritage register (ZIP → extract)
    hr_dir = ref / "heritage_register"
    if not hr_dir.exists():
        hr_zip = ref / "heritage_register.zip"
        _download(_HERITAGE_REGISTER_URL, hr_zip)
        hr_dir.mkdir()
        with zipfile.ZipFile(hr_zip) as zf:
            zf.extractall(hr_dir)
        hr_zip.unlink()

    # Heritage conservation districts (ZIP → extract)
    hd_dir = ref / "heritage_districts"
    if not hd_dir.exists():
        hd_zip = ref / "heritage_districts.zip"
        _download(_HERITAGE_DISTRICTS_URL, hd_zip)
        hd_dir.mkdir()
        with zipfile.ZipFile(hd_zip) as zf:
            zf.extractall(hd_dir)
        hd_zip.unlink()

    # Secondary plans GeoJSON
    sp_geojson = ref / "secondary_plans.geojson"
    if not sp_geojson.exists():
        _download(_SECONDARY_PLANS_URL, sp_geojson)

    # MTSA boundaries (ZIP → extract SHP, Major Transit Station Areas)
    mtsa_shp = ref / "mtsa.shp"
    if not mtsa_shp.exists():
        mtsa_dir = ref / "mtsa"
        if not mtsa_dir.exists():
            mtsa_zip = ref / "mtsa.zip"
            try:
                _download(_MTSA_URL, mtsa_zip)
                mtsa_dir.mkdir()
                with zipfile.ZipFile(mtsa_zip) as zf:
                    zf.extractall(mtsa_dir)
                mtsa_zip.unlink()
                # Find the .shp file in the extracted directory and move to root ref dir
                shp_files = list(mtsa_dir.glob("*.shp"))
                if shp_files:
                    shp_file = shp_files[0]
                    # Copy all shapefile components to ref dir
                    # (*.shp, *.shx, *.dbf, *.prj, *.cpg, etc.)
                    for ext in ["shp", "shx", "dbf", "prj", "cpg"]:
                        src = mtsa_dir / f"{shp_file.stem}.{ext}"
                        if src.exists():
                            (ref / f"mtsa.{ext}").write_bytes(src.read_bytes())
                # Clean up extracted directory
                shutil.rmtree(mtsa_dir)
            except Exception:
                logger.warning(
                    "MTSA boundaries not available — in_mtsa will be 0 for all rows"
                )

    # Ward profiles CSV (computed from two XLSX downloads)
    ward_profiles_csv = ref / "ward_profiles.csv"
    if not ward_profiles_csv.exists():
        _fetch_ward_profiles_csv(ref)

    # Section 37 community benefits CSV → parquet
    section37_parquet = ref / "section37.parquet"
    if not section37_parquet.exists():
        try:
            _fetch_section37(ref)
        except Exception:
            logger.warning(
                "Section 37 data not available — s37_monetary_value will be null"
            )

    # TRCA regulated areas GeoJSON (paginated ArcGIS FeatureServer)
    trca_geojson = ref / "trca_regulated_areas.geojson"
    if not trca_geojson.exists():
        try:
            _fetch_trca_regulated_areas(trca_geojson)
        except Exception:
            logger.warning(
                "TRCA regulated areas not available — in_trca_regulated_area will be 0"
            )

    # Ontario Greenbelt boundary GeoJSON
    greenbelt_geojson = ref / "greenbelt.geojson"
    if not greenbelt_geojson.exists():
        try:
            _fetch_greenbelt(greenbelt_geojson)
        except Exception:
            logger.warning(
                "Greenbelt boundary not available — in_greenbelt will be 0 for all rows"
            )
