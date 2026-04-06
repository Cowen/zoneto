"""Enrichment pipeline: fetch reference data, label outcomes, spatial join."""

from __future__ import annotations

import logging
import zipfile
from datetime import date as _date
from pathlib import Path

import duckdb
import joblib
import numpy as np
import polars as pl
import pyproj
from sklearn.decomposition import TruncatedSVD
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.pipeline import Pipeline as SklearnPipeline

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Reference dataset URLs
# ---------------------------------------------------------------------------
_ZONING_URL = (
    "https://ckan0.cf.opendata.inter.prod-toronto.ca"
    "/dataset/34927e44-fc11-4336-a8aa-a0dfb27658b7"
    "/resource/d75fa1ed-cd04-4a0b-bb6d-2b928ffffa6e"
    "/download/zoning-area-4326.geojson"
)
_HERITAGE_REGISTER_URL = (
    "https://ckan0.cf.opendata.inter.prod-toronto.ca"
    "/dataset/e41da515-5ad1-4bc3-85ea-18ec9e55cd33"
    "/resource/108b1080-d048-439f-a9e8-e8d6cd81bddb"
    "/download/heritage_register_address_points_wgs84.zip"
)
_HERITAGE_DISTRICTS_URL = (
    "https://ckan0.cf.opendata.inter.prod-toronto.ca"
    "/dataset/37a3c911-0813-4e87-90ed-3b9fa6156a63"
    "/resource/8e6b9347-63a8-4dac-91fb-a6491a8c1e5a"
    "/download/heritageconservationdistrict.zip"
)
_SECONDARY_PLANS_URL = (
    "https://ckan0.cf.opendata.inter.prod-toronto.ca"
    "/dataset/70a544e9-ee83-43a4-b0be-0dc973627ad7"
    "/resource/08099a8c-a598-4ca3-8395-e4159cc1ec1a"
    "/download/secondary-plans-data-2017-4326.geojson"
)
_ZONING_HEIGHT_URL = (
    "https://ckan0.cf.opendata.inter.prod-toronto.ca"
    "/dataset/34927e44-fc11-4336-a8aa-a0dfb27658b7"
    "/resource/eec27e60-7c2d-4c46-8fa1-b64f441bcc39"
    "/download/zoning-height-overlay-4326.geojson"
)
_MTSA_URL = (
    "https://ckan0.cf.opendata.inter.prod-toronto.ca"
    "/dataset/f7128f82-810d-4677-8166-8892f51969d3"
    "/resource/2b66ca26-b345-4e62-8568-dcd2cd6c3f91"
    "/download/majortransitstationareadelinations_jan2026.shp.zip"
)
_WARD_CENSUS_URL = (
    "https://ckan0.cf.opendata.inter.prod-toronto.ca"
    "/dataset/6678e1a6-d25f-4dff-b2b7-aa8f042bc2eb"
    "/resource/16a31e1d-b4d9-4cf0-b5b3-2e3937cb4121"
    "/download/2023-wardprofiles-2011-2021-censusdata_rev0719.xlsx"
)
_WARD_GEO_URL = (
    "https://ckan0.cf.opendata.inter.prod-toronto.ca"
    "/dataset/6678e1a6-d25f-4dff-b2b7-aa8f042bc2eb"
    "/resource/9398da69-0622-4eb1-a125-4f8c7f1016a4"
    "/download/2023-wardprofiles-geographicareas.xlsx"
)

# ---------------------------------------------------------------------------
# Status label sets (lowercase after strip)
# ---------------------------------------------------------------------------
_DEV_APPROVED_SET: frozenset[str] = frozenset(
    {
        "noac issued",
        "council approved",
        "draft plan approved",
        "final approval completed",
        "omb approved",
        "approved",
        "omb partially approved",
    }
)
_DEV_REFUSED_SET: frozenset[str] = frozenset({"refused", "omb refused"})
_DEV_APPEALED_SET: frozenset[str] = frozenset(
    {
        "omb appeal",
        "appeal received",
        "omb approved",
        "omb refused",
        "omb partially approved",
    }
)
_DEV_ACTIVE_SET: frozenset[str] = frozenset(
    {
        "under review",
        "on hold",
        "referred",
        "deferred",
        "information requested",
    }
)
_COA_APPROVED_SET: frozenset[str] = frozenset(
    {
        "approved",
        "conditional approval",
        "approved with conditions",
        "approved on condition",
    }
)
_COA_REFUSED_SET: frozenset[str] = frozenset({"refused"})
_DEV_SURVIVAL_TYPES: frozenset[str] = frozenset({"OZ", "SA"})
_DEV_DAYS_CAP = 3650


def _label_from_sets(
    val: str | None,
    positive_set: frozenset[str],
    negative_set: frozenset[str],
) -> int | None:
    """Map a status string to 1/0/null via two frozensets.

    Returns 1 if val (stripped, lowercased) is in positive_set,
    0 if in negative_set, None otherwise.
    """
    if val is None:
        return None
    v = val.strip().lower()
    if v in positive_set:
        return 1
    if v in negative_set:
        return 0
    return None


def _fetch_ward_profiles_csv(ref: Path) -> None:
    """Download ward profile XLSXs, compute metrics, write ward_profiles.csv.

    Output: ward_number,ward_pct_renters,ward_median_income,
    ward_pop_density,ward_pct_detached (one row per ward, wards 1–25).
    """
    import csv
    import tempfile

    import httpx
    import openpyxl

    census_xlsx = Path(tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name)
    geo_xlsx = Path(tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name)
    try:
        with httpx.Client(follow_redirects=True, timeout=120) as client:
            census_xlsx.write_bytes(
                client.get(_WARD_CENSUS_URL).raise_for_status().content
            )
            geo_xlsx.write_bytes(client.get(_WARD_GEO_URL).raise_for_status().content)

        # --- parse census XLSX ---
        wb = openpyxl.load_workbook(census_xlsx, read_only=True, data_only=True)
        ws = wb["2021 One Variable"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        # Row index 17: header (None, 'Toronto', 'Ward 1', ..., 'Ward 25')
        header = rows[17]
        # Build index: ward_number (int) → column index in rows
        ward_col_idx: dict[int, int] = {}
        for col_i, val in enumerate(header):
            if val and str(val).startswith("Ward "):
                try:
                    ward_col_idx[int(str(val).replace("Ward ", "").strip())] = col_i
                except ValueError:
                    pass

        def _cell(row_idx: int, ward_num: int) -> float | None:
            col_i = ward_col_idx.get(ward_num)
            if col_i is None:
                return None
            v = rows[row_idx][col_i]
            return float(v) if v is not None else None

        # Row 18: total population ('Total - Age', toronto_total, ward1, ...)
        # Row 43: total occupied dwellings
        # Row 44: single-detached dwellings
        # Row 1384: median household income
        # Row 1390: tenant households
        # Row 1394: owner households

        # --- parse geographic areas XLSX for ward area (sq km) ---
        wb2 = openpyxl.load_workbook(geo_xlsx, read_only=True, data_only=True)
        ws2 = wb2[wb2.sheetnames[0]]
        geo_rows = list(ws2.iter_rows(values_only=True))
        wb2.close()

        # Row 11: header; rows 12+ are (ward_num, area_sq_km, ...)
        ward_area: dict[int, float] = {}
        for geo_row in geo_rows[12:]:
            if geo_row[0] is None:
                break
            try:
                ward_area[int(geo_row[0])] = float(geo_row[1])
            except (TypeError, ValueError):
                pass

        # Write output CSV
        out_path = ref / "ward_profiles.csv"
        with out_path.open("w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "ward_number",
                    "ward_pct_renters",
                    "ward_median_income",
                    "ward_pop_density",
                    "ward_pct_detached",
                ]
            )
            for ward_num in sorted(ward_col_idx.keys()):
                population = _cell(18, ward_num)
                total_dwellings = _cell(43, ward_num)
                single_detached = _cell(44, ward_num)
                median_income = _cell(1384, ward_num)
                tenant = _cell(1390, ward_num)
                owner = _cell(1394, ward_num)

                pct_renters: float | None = None
                if tenant is not None and owner is not None and (tenant + owner) > 0:
                    pct_renters = tenant / (tenant + owner) * 100.0

                pop_density: float | None = None
                area = ward_area.get(ward_num)
                if population is not None and area is not None and area > 0:
                    pop_density = population / area

                pct_detached: float | None = None
                if (
                    single_detached is not None
                    and total_dwellings is not None
                    and total_dwellings > 0
                ):
                    pct_detached = single_detached / total_dwellings * 100.0

                writer.writerow(
                    [ward_num, pct_renters, median_income, pop_density, pct_detached]
                )
    finally:
        census_xlsx.unlink(missing_ok=True)
        geo_xlsx.unlink(missing_ok=True)


def _enrich_ward_features(df: pl.DataFrame, data_dir: Path) -> pl.DataFrame:
    """Add ward profile features (ward_pct_renters, ward_median_income, etc.) to df.

    Reads ward_profiles.csv (simple format: one row per ward).
    Left-joins on ward_number (normalizing format — dev uses "Ward N", COA uses "N").
    Returns enriched DataFrame with new columns.
    """
    ref = data_dir / "reference"
    ward_profiles_path = ref / "ward_profiles.csv"

    if not ward_profiles_path.exists():
        return df.with_columns(
            pl.lit(None, dtype=pl.Float64).alias("ward_pct_renters"),
            pl.lit(None, dtype=pl.Float64).alias("ward_median_income"),
            pl.lit(None, dtype=pl.Float64).alias("ward_pop_density"),
            pl.lit(None, dtype=pl.Float64).alias("ward_pct_detached"),
        )

    # Simple CSV: one row per ward with precomputed metric columns
    ward_df = pl.read_csv(ward_profiles_path)

    # Build lookup: ward_number (int) → metrics dict
    ward_data: dict[int, dict[str, float | None]] = {}
    for row in ward_df.iter_rows(named=True):
        try:
            w = int(row["ward_number"])
        except (TypeError, ValueError):
            continue
        ward_data[w] = {
            "ward_pct_renters": row.get("ward_pct_renters"),
            "ward_median_income": row.get("ward_median_income"),
            "ward_pop_density": row.get("ward_pop_density"),
            "ward_pct_detached": row.get("ward_pct_detached"),
        }

    def normalize_ward(ward_val: object) -> int | None:
        if ward_val is None:
            return None
        ward_str = str(ward_val).replace("Ward ", "").strip()
        try:
            return int(ward_str)
        except ValueError:
            return None

    ward_nums = [normalize_ward(w) for w in df["ward_number"]]

    def _lookup(metric: str) -> list[float | None]:
        return [
            ward_data.get(w, {}).get(metric) if w is not None else None
            for w in ward_nums
        ]

    F = pl.Float64
    return df.with_columns(
        pl.Series("ward_pct_renters", _lookup("ward_pct_renters"), dtype=F),
        pl.Series("ward_median_income", _lookup("ward_median_income"), dtype=F),
        pl.Series("ward_pop_density", _lookup("ward_pop_density"), dtype=F),
        pl.Series("ward_pct_detached", _lookup("ward_pct_detached"), dtype=F),
    )


def _download(url: str, dest: Path) -> None:
    """Download *url* to *dest* (binary)."""
    import httpx

    with httpx.Client(follow_redirects=True, timeout=120) as client:
        r = client.get(url)
        r.raise_for_status()
        dest.write_bytes(r.content)


def fetch_reference(data_dir: Path = Path("data")) -> None:
    """Download all reference datasets to *data_dir*/reference/.

    Idempotent: skips files that already exist.
    """
    ref = data_dir / "reference"
    ref.mkdir(parents=True, exist_ok=True)

    # Zoning GeoJSON (full-city, WGS84)
    zoning_geojson = ref / "zoning.geojson"
    if not zoning_geojson.exists():
        _download(_ZONING_URL, zoning_geojson)

    # Zoning height overlay GeoJSON (max storeys/height per area, WGS84)
    zoning_height = ref / "zoning_height.geojson"
    if not zoning_height.exists():
        _download(_ZONING_HEIGHT_URL, zoning_height)

    # Heritage register (ZIP → extract)
    hr_dir = ref / "heritage_register"
    if not hr_dir.exists():
        hr_zip = ref / "heritage_register.zip"
        _download(_HERITAGE_REGISTER_URL, hr_zip)
        hr_dir.mkdir()
        with zipfile.ZipFile(hr_zip) as zf:
            zf.extractall(hr_dir)
        hr_zip.unlink()

    # Heritage conservation districts (ZIP → extract)
    hd_dir = ref / "heritage_districts"
    if not hd_dir.exists():
        hd_zip = ref / "heritage_districts.zip"
        _download(_HERITAGE_DISTRICTS_URL, hd_zip)
        hd_dir.mkdir()
        with zipfile.ZipFile(hd_zip) as zf:
            zf.extractall(hd_dir)
        hd_zip.unlink()

    # Secondary plans GeoJSON
    sp_geojson = ref / "secondary_plans.geojson"
    if not sp_geojson.exists():
        _download(_SECONDARY_PLANS_URL, sp_geojson)

    # MTSA boundaries (ZIP → extract SHP, Major Transit Station Areas)
    mtsa_shp = ref / "mtsa.shp"
    if not mtsa_shp.exists():
        mtsa_dir = ref / "mtsa"
        if not mtsa_dir.exists():
            mtsa_zip = ref / "mtsa.zip"
            try:
                _download(_MTSA_URL, mtsa_zip)
                mtsa_dir.mkdir()
                with zipfile.ZipFile(mtsa_zip) as zf:
                    zf.extractall(mtsa_dir)
                mtsa_zip.unlink()
                # Find the .shp file in the extracted directory and move to root ref dir
                shp_files = list(mtsa_dir.glob("*.shp"))
                if shp_files:
                    shp_file = shp_files[0]
                    # Copy all shapefile components to ref dir
                    # (*.shp, *.shx, *.dbf, *.prj, *.cpg, etc.)
                    for ext in ["shp", "shx", "dbf", "prj", "cpg"]:
                        src = mtsa_dir / f"{shp_file.stem}.{ext}"
                        if src.exists():
                            (ref / f"mtsa.{ext}").write_bytes(src.read_bytes())
                # Clean up extracted directory
                import shutil

                shutil.rmtree(mtsa_dir)
            except Exception:
                logger.warning(
                    "MTSA boundaries not available — in_mtsa will be 0 for all rows"
                )

    # Ward profiles CSV (computed from two XLSX downloads)
    ward_profiles_csv = ref / "ward_profiles.csv"
    if not ward_profiles_csv.exists():
        _fetch_ward_profiles_csv(ref)


def enrich_coa(data_dir: Path = Path("data")) -> int:
    """Enrich COA parquet with outcome labels; write data/enriched/coa.parquet.

    Returns row count written.
    """
    df = pl.read_parquet(data_dir / "coa", hive_partitioning=True)

    # Deduplicate on application key (consolidated CSV may overlap individual year CSVs)
    if "reference_file" in df.columns:
        df = df.unique(subset=["reference_file"], keep="first", maintain_order=True)

    # Rename ward → ward_number (cast to str for consistency)
    df = df.rename({"ward": "ward_number"}).with_columns(
        pl.col("ward_number").cast(pl.Utf8)
    )

    # year_submitted from in_date
    df = df.with_columns(
        pl.col("in_date").dt.year().cast(pl.Int32).alias("year_submitted")
    )

    # coa_approved label
    df = df.with_columns(
        pl.col("c_of_a_descision")
        .map_elements(
            lambda v: _label_from_sets(v, _COA_APPROVED_SET, _COA_REFUSED_SET),
            return_dtype=pl.Int8,
        )
        .alias("coa_approved")
    )

    # coa_days_to_approval — only for approved rows with both dates present.
    # Cap at 730 days (2 years): outliers beyond this are almost certainly data
    # errors (legacy applications closed years after filing) that destabilize
    # the regression across CV folds.
    _CAP_DAYS = 730
    days = (pl.col("finaldate") - pl.col("in_date")).dt.total_days().cast(pl.Int32)
    df = df.with_columns(
        pl.when((pl.col("coa_approved") == 1) & (days <= _CAP_DAYS))
        .then(days)
        .otherwise(None)
        .alias("coa_days_to_approval")
    )

    # Enrich with ward profiles
    df = _enrich_ward_features(df, data_dir)

    out = data_dir / "enriched"
    out.mkdir(parents=True, exist_ok=True)
    df.write_parquet(out / "coa.parquet")
    return len(df)


def _spatial_join_dev(df: pl.DataFrame, data_dir: Path) -> pl.DataFrame:
    """Add zoning_class, secondary_plan_name, in_heritage_register,
    in_heritage_district, in_secondary_plan, in_mtsa columns via DuckDB spatial join.

    Rows with null or garbage x/y get null/0 enrichment values.
    """
    ref = data_dir / "reference"

    # Reproject x/y from EPSG:2952 (NAD83 / MTM Zone 10, City of Toronto internal CRS)
    # to EPSG:4326. Median dev application x ≈ 313,000 (MTM false easting 304,800),
    # confirming EPSG:2952 — not EPSG:26917 (UTM 17N, false easting 500,000) which
    # would map Toronto parcels to Michigan.
    transformer = pyproj.Transformer.from_crs("EPSG:2952", "EPSG:4326", always_xy=True)
    xs = df["x"].cast(pl.Float64, strict=False).to_list()
    ys = df["y"].cast(pl.Float64, strict=False).to_list()

    lons: list[float | None] = []
    lats: list[float | None] = []
    for x_val, y_val in zip(xs, ys):
        if x_val is None or y_val is None or y_val < 4_000_000:
            lons.append(None)
            lats.append(None)
        else:
            lon, lat = transformer.transform(x_val, y_val)
            lons.append(lon)
            lats.append(lat)

    df = df.with_columns(
        pl.Series("lon", lons, dtype=pl.Float64),
        pl.Series("lat", lats, dtype=pl.Float64),
        pl.Series("_rid", range(len(df)), dtype=pl.Int64),
    )

    con = duckdb.connect()
    con.execute("INSTALL spatial; LOAD spatial;")

    con.register("apps", df.to_arrow())

    zoning_geojson_path = str(ref / "zoning.geojson").replace("'", "''")
    hr_shp = str(next((ref / "heritage_register").glob("*.shp"))).replace(
        "'",
        "''",
    )
    hd_shp = str(next((ref / "heritage_districts").glob("*.shp"))).replace(
        "'",
        "''",
    )
    sp_geojson = str(ref / "secondary_plans.geojson").replace("'", "''")

    result = con.execute(f"""
        WITH pts AS (
            SELECT
                _rid,
                lon, lat,
                CASE WHEN lon IS NOT NULL AND lat IS NOT NULL
                     THEN ST_Point(lon, lat) END AS geom
            FROM apps
        ),
        zoning_join AS (
            SELECT DISTINCT ON (p._rid)
                p._rid,
                z.ZN_ZONE AS zoning_class,
                z.UNITS AS zoning_max_units,
                z.DENSITY AS zoning_max_density
            FROM pts p
            LEFT JOIN ST_Read('{zoning_geojson_path}') z
                ON ST_Within(p.geom, z.geom)
            WHERE p.geom IS NOT NULL
        ),
        hr_join AS (
            SELECT DISTINCT ON (p._rid)
                p._rid,
                1::TINYINT AS in_heritage_register
            FROM pts p
            JOIN ST_Read('{hr_shp}') h
                ON ST_Intersects(ST_Buffer(p.geom, 0.0002), h.geom)
            WHERE p.geom IS NOT NULL
        ),
        hd_join AS (
            SELECT DISTINCT ON (p._rid)
                p._rid,
                1::TINYINT AS in_heritage_district
            FROM pts p
            JOIN ST_Read('{hd_shp}') h
                ON ST_Within(p.geom, h.geom)
            WHERE p.geom IS NOT NULL
        ),
        sp_join AS (
            SELECT DISTINCT ON (p._rid)
                p._rid,
                s.SECONDARY_PLAN_NAME AS secondary_plan_name
            FROM pts p
            JOIN ST_Read('{sp_geojson}') s
                ON ST_Within(p.geom, s.geom)
            WHERE p.geom IS NOT NULL
        )
        SELECT
            a.*,
            z.zoning_class AS zoning_class,
            z.zoning_max_units AS zoning_max_units,
            z.zoning_max_density AS zoning_max_density,
            COALESCE(h.in_heritage_register, 0)
                AS in_heritage_register,
            COALESCE(d.in_heritage_district, 0)
                AS in_heritage_district,
            sp.secondary_plan_name AS secondary_plan_name,
            CASE WHEN sp.secondary_plan_name IS NOT NULL
                 THEN 1 ELSE 0 END AS in_secondary_plan
        FROM apps a
        LEFT JOIN zoning_join z ON a._rid = z._rid
        LEFT JOIN hr_join h ON a._rid = h._rid
        LEFT JOIN hd_join d ON a._rid = d._rid
        LEFT JOIN sp_join sp ON a._rid = sp._rid
    """).pl()

    con.close()

    # Cast zoning limit columns to expected types (DuckDB may return bigint/double)
    result = result.with_columns(
        pl.col("zoning_max_units").cast(pl.Int32, strict=False),
        pl.col("zoning_max_density").cast(pl.Float64, strict=False),
    )

    # Add zoning max storeys from height overlay
    height_geojson = ref / "zoning_height.geojson"
    result = _add_height_feature(result, height_geojson)

    # Add MTSA feature using the downloaded boundary shapefile
    mtsa_shp = ref / "mtsa.shp"
    result = _add_mtsa_feature(result, mtsa_shp)

    return result


def _add_height_feature(
    df: pl.DataFrame, height_path: Path
) -> pl.DataFrame:
    """Add zoning_max_storeys (Int32) via DuckDB spatial join against height overlay.

    Per the by-law data dictionary, HT_STORIES <= 0 means "no limit" — treated
    as null. Rows with null lat/lon or outside height overlay get null.
    """
    if (
        not height_path.exists()
        or "lat" not in df.columns
        or "lon" not in df.columns
    ):
        return df.with_columns(
            pl.lit(None, dtype=pl.Int32).alias("zoning_max_storeys")
        )

    valid_mask = df["lat"].is_not_null() & df["lon"].is_not_null()
    valid_df = df.filter(valid_mask)

    if len(valid_df) == 0:
        return df.with_columns(
            pl.lit(None, dtype=pl.Int32).alias("zoning_max_storeys")
        )

    escaped = str(height_path).replace("'", "''")
    con = duckdb.connect()
    con.execute("INSTALL spatial; LOAD spatial;")
    con.register("apps", valid_df.to_arrow())

    result = con.execute(f"""
        SELECT DISTINCT ON (apps._rid)
            apps._rid,
            CASE WHEN ht.HT_STORIES > 0
                 THEN ht.HT_STORIES ELSE NULL END
                 AS zoning_max_storeys
        FROM apps
        LEFT JOIN ST_Read('{escaped}') ht
            ON ST_Within(ST_Point(apps.lon, apps.lat), ht.geom)
    """).pl()

    con.close()

    rid_to_ht: dict[int, int | None] = dict(
        zip(
            result["_rid"].to_list(),
            result["zoning_max_storeys"]
            .cast(pl.Int32, strict=False)
            .to_list(),
        )
    )
    all_rids = (
        df["_rid"].to_list() if "_rid" in df.columns else list(range(len(df)))
    )
    ht_series = pl.Series(
        "zoning_max_storeys",
        [rid_to_ht.get(rid) for rid in all_rids],
        dtype=pl.Int32,
    )
    return df.with_columns(ht_series)


def _add_mtsa_feature(df: pl.DataFrame, mtsa_path: Path) -> pl.DataFrame:
    """Add in_mtsa (Int8) column via DuckDB spatial join against MTSA boundaries.

    Rows with null lat/lon or outside MTSA boundaries get in_mtsa=0.
    If mtsa_path does not exist, all rows get in_mtsa=0.
    mtsa_path can be any format that DuckDB ST_Read understands (.shp, .geojson, etc.).
    """
    if not mtsa_path.exists() or "lat" not in df.columns or "lon" not in df.columns:
        return df.with_columns(pl.lit(0, dtype=pl.Int8).alias("in_mtsa"))

    valid_mask = df["lat"].is_not_null() & df["lon"].is_not_null()
    valid_df = df.filter(valid_mask)

    if len(valid_df) == 0:
        return df.with_columns(pl.lit(0, dtype=pl.Int8).alias("in_mtsa"))

    mtsa_path_escaped = str(mtsa_path).replace("'", "''")
    con = duckdb.connect()
    con.execute("INSTALL spatial; LOAD spatial;")
    con.register("apps", valid_df.to_arrow())

    result = con.execute(f"""
        SELECT apps._rid,
               CASE WHEN COUNT(mtsa.geom) > 0 THEN 1 ELSE 0 END AS in_mtsa
        FROM apps
        LEFT JOIN ST_Read('{mtsa_path_escaped}') mtsa
            ON ST_Within(
                ST_Point(apps.lon, apps.lat),
                mtsa.geom
            )
        GROUP BY apps._rid
    """).pl()

    con.close()

    rid_to_mtsa: dict[int, int] = dict(
        zip(result["_rid"].to_list(), result["in_mtsa"].cast(pl.Int8).to_list())
    )
    all_rids = df["_rid"].to_list() if "_rid" in df.columns else list(range(len(df)))
    in_mtsa_series = pl.Series(
        "in_mtsa",
        [rid_to_mtsa.get(rid, 0) for rid in all_rids],
        dtype=pl.Int8,
    )
    return df.with_columns(in_mtsa_series)


def _extract_text_features(
    df: pl.DataFrame,
    model_dir: Path,
    *,
    n_components: int = 20,
) -> tuple[pl.DataFrame, SklearnPipeline]:
    """Extract TF-IDF + TruncatedSVD features from the description column.

    Fits TfidfVectorizer (max_features=5000) + TruncatedSVD on the description
    column. Serializes the pipeline to model_dir/desc_tfidf.joblib.
    Adds desc_svd_0..{n_components-1} columns to the DataFrame.

    Rows with null descriptions are treated as empty strings (→ zero SVD vector).
    Returns (enriched_df, fitted_pipeline).
    """
    svd_col_names = [f"desc_svd_{i}" for i in range(n_components)]
    zero_svd = [pl.lit(0.0, dtype=pl.Float64).alias(col) for col in svd_col_names]

    if "description" not in df.columns:
        pipeline = SklearnPipeline(
            [
                (
                    "tfidf",
                    TfidfVectorizer(
                        max_features=5000,
                        ngram_range=(1, 2),
                        min_df=1,
                        sublinear_tf=True,
                    ),
                ),
                ("svd", TruncatedSVD(n_components=n_components, random_state=42)),
            ]
        )
        return df.with_columns(zero_svd), pipeline

    texts = df["description"].fill_null("").cast(pl.String).to_list()

    # Fit TF-IDF first to know vocabulary size
    tfidf = TfidfVectorizer(
        max_features=5000, ngram_range=(1, 2), min_df=1, sublinear_tf=True
    )
    tfidf_matrix = tfidf.fit_transform(texts)
    n_features = tfidf_matrix.shape[1]

    # TruncatedSVD requires n_components < n_features (at least 2 features needed)
    # If corpus is too small to fit SVD, return zero vectors
    if n_features < 2:
        logger.warning(
            "TF-IDF vocabulary too small (%d features) — "
            "desc_svd columns will be zeros",
            n_features,
        )
        vectors = np.zeros((len(texts), n_components))
        # Return zero vectors without serializing an unfitted pipeline
        # score.py will fall back to enriched parquet columns if pipeline is absent
        pipeline = SklearnPipeline(
            [
                (
                    "tfidf",
                    TfidfVectorizer(
                        max_features=5000,
                        ngram_range=(1, 2),
                        min_df=1,
                        sublinear_tf=True,
                    ),
                ),
                ("svd", TruncatedSVD(n_components=n_components, random_state=42)),
            ]
        )
        # Don't serialize — unfitted pipeline
        return df.with_columns(zero_svd), pipeline

    # Fit SVD with safe number of components
    safe_n = min(n_components, n_features - 1)
    svd = TruncatedSVD(n_components=safe_n, random_state=42)
    vectors = svd.fit_transform(tfidf_matrix)
    # Pad with zeros if SVD used fewer components than requested
    if vectors.shape[1] < n_components:
        pad = np.zeros((vectors.shape[0], n_components - vectors.shape[1]))
        vectors = np.hstack([vectors, pad])

    # Only serialize when we have a properly fitted pipeline
    pipeline = SklearnPipeline(
        [
            ("tfidf", tfidf),
            ("svd", svd),
        ]
    )

    model_dir.mkdir(parents=True, exist_ok=True)
    joblib.dump(pipeline, model_dir / "desc_tfidf.joblib")

    svd_cols = [
        pl.Series(f"desc_svd_{i}", vectors[:, i].tolist(), dtype=pl.Float64)
        for i in range(n_components)
    ]
    return df.with_columns(svd_cols), pipeline


def _compute_ward_appeal_rate_3y(df: pl.DataFrame) -> pl.DataFrame:
    """Add ward_appeal_rate_3y: rolling 3-year appeal rate per ward.

    For each row, computes the appeal rate in the same ward using only
    OZ/SA rows from the 3 years strictly before the row's year_submitted.
    Returns null when no prior data exists for the ward.
    """
    # Build per-ward-per-year appeal stats from labeled OZ/SA rows
    labeled = df.filter(pl.col("dev_appealed").is_not_null()).select(
        ["ward_number", "year_submitted", "dev_appealed"]
    )
    if labeled.is_empty():
        return df.with_columns(
            pl.lit(None, dtype=pl.Float64).alias("ward_appeal_rate_3y")
        )

    ward_year_stats = labeled.group_by(["ward_number", "year_submitted"]).agg(
        pl.col("dev_appealed").sum().alias("appeals"),
        pl.col("dev_appealed").count().alias("total"),
    )

    # For each unique (ward, year) in the data, compute rolling 3-year rate
    unique_ward_years = df.select(["ward_number", "year_submitted"]).unique()
    # Cross-join with ward_year_stats and filter to prior 3 years
    rates = (
        unique_ward_years.join(ward_year_stats, on="ward_number", suffix="_hist")
        .filter(
            (pl.col("year_submitted_hist") < pl.col("year_submitted"))
            & (pl.col("year_submitted_hist") >= pl.col("year_submitted") - 3)
        )
        .group_by(["ward_number", "year_submitted"])
        .agg(
            (pl.col("appeals").sum() / pl.col("total").sum()).alias(
                "ward_appeal_rate_3y"
            ),
        )
    )

    return df.join(rates, on=["ward_number", "year_submitted"], how="left")


def enrich_dev(data_dir: Path = Path("data")) -> int:
    """Enrich dev_applications parquet with spatial features and outcome labels.

    Writes data/enriched/dev_applications.parquet. Returns row count.
    """
    df = pl.read_parquet(data_dir / "dev_applications", hive_partitioning=True)

    # --- AIC application records: prefer over CKAN for matching folderrsn ---
    # If data/aic_applications/ exists (from 'zoneto aic --full'), merge it:
    # - AIC records override CKAN records for the same folderrsn
    # - AIC-only records (not in CKAN) are added to the dataset
    aic_apps_path = data_dir / "aic_applications"
    if aic_apps_path.exists() and any(aic_apps_path.rglob("*.parquet")):
        aic_df = pl.read_parquet(aic_apps_path, hive_partitioning=True)
        # Ensure folderrsn is String for join
        if "folderrsn" in aic_df.columns:
            aic_df = aic_df.with_columns(pl.col("folderrsn").cast(pl.String))
        if "folderrsn" in df.columns:
            df = df.with_columns(pl.col("folderrsn").cast(pl.String))
        # Align schemas: use CKAN schema as base, update matching rows from AIC
        aic_rsns = (
            set(aic_df["folderrsn"].to_list())
            if "folderrsn" in aic_df.columns
            else set()
        )
        # Keep CKAN rows not in AIC; replace CKAN rows that AIC has; add AIC-only rows
        ckan_only = df.filter(~pl.col("folderrsn").is_in(list(aic_rsns)))
        # For AIC records, map column names to CKAN schema where possible
        shared_cols = [c for c in df.columns if c in aic_df.columns]
        aic_aligned = aic_df.select(shared_cols)
        # Add missing CKAN columns as nulls
        for col in df.columns:
            if col not in aic_aligned.columns:
                aic_aligned = aic_aligned.with_columns(
                    pl.lit(None).cast(df[col].dtype).alias(col)
                )
        aic_aligned = aic_aligned.select(df.columns)
        df = pl.concat([ckan_only, aic_aligned], how="diagonal")
        logger.info(
            "enrich_dev: merged AIC records — %d CKAN-only, %d from AIC (%d total)",
            len(ckan_only),
            len(aic_aligned),
            len(df),
        )

    # year_submitted and _submitted_date from date_submitted
    # (may be Date/Datetime or String like "2022-08-09T00:00:00")
    _ds_dtype = df["date_submitted"].dtype
    if _ds_dtype.is_temporal():
        _year_expr = pl.col("date_submitted").dt.year().cast(pl.Int32)
        _date_expr = pl.col("date_submitted").cast(pl.Date)
    else:
        # String like "2022-08-09T00:00:00" — slice to date part
        _year_expr = (
            pl.col("date_submitted").str.slice(0, 4).cast(pl.Int32, strict=False)
        )
        _date_expr = pl.col("date_submitted").str.slice(0, 10).str.to_date(strict=False)
    df = df.with_columns(
        _year_expr.alias("year_submitted"),
        _date_expr.alias("_submitted_date"),
    )

    # has_community_meeting
    df = df.with_columns(
        pl.col("community_meeting_date")
        .is_not_null()
        .cast(pl.Int8)
        .alias("has_community_meeting"),
    )

    # Spatial enrichment (monkeypatchable)
    df = _spatial_join_dev(df, data_dir)

    df = df.with_columns(
        pl.col("status")
        .map_elements(
            lambda v: _label_from_sets(v, _DEV_APPROVED_SET, _DEV_REFUSED_SET),
            return_dtype=pl.Int8,
        )
        .alias("dev_approved"),
        # dev_appealed: 1=appeal filed, 0=closed without appeal, null=active/non-OZ/SA.
        # Restricted to OZ+SA only — these are the types with AIC decision milestones.
        # Covers ALL closed OZ/SA applications (not just explicitly-approved ones) so
        # the training base rate reflects the true Toronto appeal rate (~15-25%),
        # not a selection-biased 50/50 split.
        pl.when(~pl.col("application_type").is_in(list(_DEV_SURVIVAL_TYPES)))
        .then(None)
        .when(pl.col("status").is_null())
        .then(None)
        .when(
            pl.col("status")
            .str.strip_chars()
            .str.to_lowercase()
            .is_in(list(_DEV_ACTIVE_SET))
        )
        .then(None)
        .when(
            pl.col("status")
            .str.strip_chars()
            .str.to_lowercase()
            .is_in(list(_DEV_APPEALED_SET))
        )
        .then(pl.lit(1, dtype=pl.Int8))
        .otherwise(pl.lit(0, dtype=pl.Int8))
        .alias("dev_appealed"),
    )

    # is_active: 1 for pending applications (not an ML feature — output flag only)
    df = df.with_columns(
        pl.col("status")
        .map_elements(
            lambda v: (
                1 if (v is not None and v.strip().lower() in _DEV_ACTIVE_SET) else 0
            ),
            return_dtype=pl.Int8,
        )
        .alias("is_active")
    )

    # ward_appeal_rate_3y: rolling 3-year appeal rate per ward using only prior years.
    # Avoids temporal leakage — each row only sees data from years before its own.
    # Uses OZ+SA rows with non-null dev_appealed for the base rate calculation.
    df = _compute_ward_appeal_rate_3y(df)

    # has_parent_application: SA/CD linked to a parent OZ implies upstream rezoning
    # decided
    if "parent_folder_number" in df.columns:
        df = df.with_columns(
            pl.col("parent_folder_number")
            .is_not_null()
            .cast(pl.Int8)
            .alias("has_parent_application")
        )
    else:
        df = df.with_columns(pl.lit(0, dtype=pl.Int8).alias("has_parent_application"))

    # postal_fsa: neighbourhood proxy (first 3 chars of postal code e.g. "M5V")
    if "postal" in df.columns:
        df = df.with_columns(pl.col("postal").str.slice(0, 3).alias("postal_fsa"))
    else:
        df = df.with_columns(pl.lit(None, dtype=pl.String).alias("postal_fsa"))

    # Enrich with ward profiles
    df = _enrich_ward_features(df, data_dir)

    # --- AIC decision date join and survival labels ---
    aic_path = data_dir / "reference" / "aic_decisions.parquet"
    if aic_path.exists() and "folderrsn" in df.columns:
        aic = pl.read_parquet(aic_path).select(["folderrsn", "decision_date"])
        df = df.join(aic, on="folderrsn", how="left")
    else:
        df = df.with_columns(pl.lit(None, dtype=pl.Date).alias("decision_date"))

    _today = _date.today()

    # dev_days_to_decision: days from date_submitted to decision_date (OZ/SA only)
    # Intermediate column _raw_days used to simplify capping logic.
    df = df.with_columns(
        pl.when(
            pl.col("application_type").is_in(list(_DEV_SURVIVAL_TYPES))
            & pl.col("decision_date").is_not_null()
        )
        .then(
            (pl.col("decision_date") - pl.col("_submitted_date"))
            .dt.total_days()
            .cast(pl.Int32)
        )
        .otherwise(None)
        .alias("_raw_days")
    )
    df = df.with_columns(
        pl.when(
            pl.col("application_type").is_in(list(_DEV_SURVIVAL_TYPES))
            & pl.col("decision_date").is_not_null()
            & (pl.col("_raw_days") <= _DEV_DAYS_CAP)
        )
        .then(pl.col("_raw_days"))
        .otherwise(None)
        .cast(pl.Int32)
        .alias("dev_days_to_decision")
    )

    # dev_decision_event: 1 = has decision, 0 = active, null = not OZ/SA
    df = df.with_columns(
        pl.when(~pl.col("application_type").is_in(list(_DEV_SURVIVAL_TYPES)))
        .then(None)
        .when(pl.col("decision_date").is_not_null())
        .then(pl.lit(1, dtype=pl.Int8))
        .when(pl.col("is_active") == 1)
        .then(pl.lit(0, dtype=pl.Int8))
        .otherwise(None)
        .alias("dev_decision_event")
    )

    # dev_days_observed: actual days to decision for events (uncapped, for survival
    # model time axis); today-submitted for censored; null for non-OZ/SA.
    df = df.with_columns(
        pl.when(pl.col("dev_decision_event") == 1)
        .then(pl.col("_raw_days"))  # use uncapped actual time for survival model
        .when(pl.col("dev_decision_event") == 0)
        .then(
            (pl.lit(_today) - pl.col("_submitted_date")).dt.total_days().cast(pl.Int32)
        )
        .otherwise(None)
        .alias("dev_days_observed")
    )
    df = df.drop("_raw_days", "_submitted_date")

    # proposed_storeys: extract storey count from description (e.g. "12 storey",
    # "28-storey", "3 storeys") — regex captures first match, case-insensitive.
    if "description" in df.columns:
        df = df.with_columns(
            pl.col("description")
            .str.extract(r"(?i)(\d+)\s*-?\s*store?ys?", 1)
            .cast(pl.Int32, strict=False)
            .alias("proposed_storeys")
        )
        # proposed_units: extract unit count (e.g. "551 units", "186 dwelling units")
        df = df.with_columns(
            pl.col("description")
            .str.extract(r"(?i)(\d+)\s+(?:dwelling\s+)?units?", 1)
            .cast(pl.Int32, strict=False)
            .alias("proposed_units")
        )
    else:
        df = df.with_columns(
            pl.lit(None, dtype=pl.Int32).alias("proposed_storeys"),
            pl.lit(None, dtype=pl.Int32).alias("proposed_units"),
        )

    # unit_excess_ratio: proposed_units / zoning_max_units.
    # Values > 1.0 mean the proposal exceeds zoning — strong appeal signal.
    df = df.with_columns(
        pl.when(
            pl.col("proposed_units").is_not_null()
            & pl.col("zoning_max_units").is_not_null()
            & (pl.col("zoning_max_units") > 0)
        )
        .then(pl.col("proposed_units") / pl.col("zoning_max_units"))
        .otherwise(None)
        .cast(pl.Float64)
        .alias("unit_excess_ratio")
    )

    # storey_excess_ratio: proposed_storeys / zoning_max_storeys.
    # Per by-law readme, HT_STORIES <= 0 means "no limit" (already nulled
    # in _add_height_feature). Better coverage than unit_excess_ratio.
    df = df.with_columns(
        pl.when(
            pl.col("proposed_storeys").is_not_null()
            & pl.col("zoning_max_storeys").is_not_null()
            & (pl.col("zoning_max_storeys") > 0)
        )
        .then(pl.col("proposed_storeys") / pl.col("zoning_max_storeys"))
        .otherwise(None)
        .cast(pl.Float64)
        .alias("storey_excess_ratio")
    )

    # is_combined_application: OZ with OPA in description field (case-insensitive)
    if "description" in df.columns:
        df = df.with_columns(
            pl.when(
                (pl.col("application_type") == "OZ")
                & pl.col("description").str.to_uppercase().str.contains("OPA")
            )
            .then(pl.lit(1, dtype=pl.Int8))
            .otherwise(pl.lit(0, dtype=pl.Int8))
            .alias("is_combined_application")
        )
    else:
        df = df.with_columns(pl.lit(0, dtype=pl.Int8).alias("is_combined_application"))

    # --- NLP features: TF-IDF + SVD on description column ---
    # NOTE: TF-IDF vocabulary is fit on the full corpus (all applications).
    # This introduces minor vocabulary leakage into temporal CV folds — the IDF
    # weights reflect future documents. The impact is small given max_features=5000
    # cap and SVD compression to 20 dimensions. Future work: fit TF-IDF per year
    # to eliminate leakage entirely, but current approach is acceptable for v1.
    _model_dir = data_dir.parent / "models"
    df, _ = _extract_text_features(df, _model_dir)

    out = data_dir / "enriched"
    out.mkdir(parents=True, exist_ok=True)
    df.write_parquet(out / "dev_applications.parquet")
    return len(df)


def enrich_permits(data_dir: Path = Path("data")) -> int:
    """Enrich permits_cleared parquet with issuance timeline label.

    Computes permit_issuance_days = issued_date - application_date (calendar days).
    Rows with null issued_date or application_date get null permit_issuance_days.
    Rows where issuance_days <= 0 are dropped (data quality).

    Writes data/enriched/permits_cleared.parquet. Returns row count written.
    """
    df = pl.read_parquet(data_dir / "permits_cleared", hive_partitioning=True)

    # Coerce string numeric columns to Float64 (remove comma thousands separators)
    _str_num_cols = ["est_const_cost", "dwelling_units_created", "dwelling_units_lost"]
    for _col in _str_num_cols:
        if _col in df.columns and df[_col].dtype == pl.Utf8:
            df = df.with_columns(
                pl.col(_col).str.replace_all(",", "").cast(pl.Float64, strict=False)
            )

    # application_year captures temporal queue-depth signal (permit office staffing
    # and backlog vary strongly by year — COVID slowdowns, policy changes, etc.)
    df = df.with_columns(
        pl.col("application_date").dt.year().cast(pl.Int32).alias("application_year")
    )

    days = (
        (pl.col("issued_date") - pl.col("application_date"))
        .dt.total_days()
        .cast(pl.Int32)
    )
    has_both = (
        pl.col("issued_date").is_not_null() & pl.col("application_date").is_not_null()
    )
    df = df.with_columns(
        pl.when(has_both).then(days).otherwise(None).alias("permit_issuance_days")
    )

    # Drop rows where issuance days is computed but non-positive (bad data)
    df = df.filter(
        pl.col("permit_issuance_days").is_null() | (pl.col("permit_issuance_days") > 0)
    )

    out = data_dir / "enriched"
    out.mkdir(parents=True, exist_ok=True)
    df.write_parquet(out / "permits_cleared.parquet")
    return len(df)


def match_olt_to_dev(
    dev_df: pl.DataFrame,
    data_dir: Path,
    *,
    confidence_threshold: float = 0.75,
) -> pl.DataFrame:
    """Fuzzy-match OLT decisions to dev_applications via address similarity.

    Reads data_dir/reference/olt_decisions.parquet (if present).
    For each dev_application, finds the OLT case with the highest address
    similarity score above `confidence_threshold`.

    Address similarity uses difflib.SequenceMatcher on normalized address
    strings (street_num + street_name vs OLT address).

    Adds columns: olt_case_number (String|null), olt_outcome (String|null),
    olt_decision_date (Date|null).

    Returns enriched DataFrame. If olt_decisions.parquet is absent,
    adds the three columns as all-null and returns unchanged.
    """
    import difflib  # stdlib

    null_cols = [
        pl.lit(None, dtype=pl.String).alias("olt_case_number"),
        pl.lit(None, dtype=pl.String).alias("olt_outcome"),
        pl.lit(None, dtype=pl.Date).alias("olt_decision_date"),
    ]

    olt_path = data_dir / "reference" / "olt_decisions.parquet"
    if not olt_path.exists():
        return dev_df.with_columns(null_cols)

    olt_df = pl.read_parquet(olt_path)
    olt_addresses = [
        a.lower().split(",")[0].strip()
        for a in olt_df["address"].fill_null("").to_list()
    ]
    olt_cases = olt_df["case_number"].fill_null("").to_list()
    olt_outcomes = olt_df["outcome"].fill_null("").to_list()
    # olt_dates are pl.Date objects; None values handled by fill_null with a sentinel
    olt_dates = olt_df["decision_date"].to_list()

    # Pre-index OLT addresses by street number (first token) for O(1) lookup.
    # Maps street_num -> list of indices into olt_addresses where
    # that street_num appears.
    olt_by_street: dict[str, list[int]] = {}
    for i, addr in enumerate(olt_addresses):
        street_num_token = addr.split()[0] if addr else ""
        if street_num_token:
            if street_num_token not in olt_by_street:
                olt_by_street[street_num_token] = []
            olt_by_street[street_num_token].append(i)

    # Fallback list: all indices when no street number match exists
    all_indices = list(range(len(olt_addresses)))

    matched_cases: list[str | None] = []
    matched_outcomes: list[str | None] = []
    matched_dates: list[object | None] = []  # will be date | None

    for row in dev_df.iter_rows(named=True):
        street_num = str(row.get("street_num") or "").strip()
        street_name = str(row.get("street_name") or "").strip()
        dev_address = f"{street_num} {street_name}".lower().strip()

        if not dev_address:
            matched_cases.append(None)
            matched_outcomes.append(None)
            matched_dates.append(None)
            continue

        # Determine which OLT addresses to compare against.
        # Prefer matching by street number; fall back to all addresses if no match.
        street_num_token = street_num if street_num else ""
        candidate_indices = olt_by_street.get(street_num_token, all_indices)

        best_ratio = 0.0
        best_idx = -1
        for i in candidate_indices:
            ratio = difflib.SequenceMatcher(None, dev_address, olt_addresses[i]).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_idx = i

        if best_ratio >= confidence_threshold and best_idx >= 0:
            matched_cases.append(olt_cases[best_idx])
            matched_outcomes.append(olt_outcomes[best_idx])
            matched_dates.append(olt_dates[best_idx])
        else:
            matched_cases.append(None)
            matched_outcomes.append(None)
            matched_dates.append(None)

    return dev_df.with_columns(
        [
            pl.Series("olt_case_number", matched_cases, dtype=pl.String),
            pl.Series("olt_outcome", matched_outcomes, dtype=pl.String),
            pl.Series("olt_decision_date", matched_dates, dtype=pl.Date),
        ]
    )
